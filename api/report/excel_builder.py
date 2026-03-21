"""
Excel builder — creates a multi-sheet workbook from the unified schema.
Sheets: Overview | Competition | Search Terms | Products | Topic Impact | Returns | Analysis
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL= PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
POSITIVE_FILL = PatternFill("solid", fgColor="E2EFDA")
NEGATIVE_FILL = PatternFill("solid", fgColor="FCE4D6")
WARN_FILL     = PatternFill("solid", fgColor="FFF2CC")

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
BOLD          = Font(bold=True)
THIN          = Side(style="thin", color="CCCCCC")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _h(ws, row, col, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill or HEADER_FILL
    cell.font = font or HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def _c(ws, row, col, value, fill=None, bold=False, align="left", fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bold:
        cell.font = Font(bold=True)
    if fmt:
        cell.number_format = fmt
    return cell


def _autofit(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min_w, min(length + 4, max_w))


def _pct(val):
    if val is None:
        return None
    try:
        return f"{float(val) * 100:.2f}%"
    except (TypeError, ValueError):
        return str(val)


def build_excel(schema: dict, markdown_report: str = "") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_overview(wb, schema)
    _sheet_competition(wb, schema)
    _sheet_search_terms(wb, schema)
    _sheet_products(wb, schema)
    _sheet_topic_impact(wb, schema)
    _sheet_returns(wb, schema)
    if markdown_report:
        _sheet_analysis(wb, markdown_report)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Sheet 1: Overview ──────────────────────────────────────────────────────────

def _sheet_overview(wb, schema):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    # Collect overview from best available source
    ov = {}
    for src in ["niche_overview", "chrome_ext", "ai_text"]:
        candidate = schema.get(src, {})
        if isinstance(candidate, dict):
            o = candidate.get("overview") or {}
            if isinstance(o, dict) and o:
                ov = o
                break

    niche = schema.get("niche_name") or ""
    for src in ["niche_overview", "products_csv", "search_terms_csv", "chrome_ext", "ai_text"]:
        s = schema.get(src, {})
        if isinstance(s, dict) and s.get("niche_name"):
            niche = s["niche_name"]
            break

    rows = [
        ("Ниша", niche),
        ("Объём поиска (360д)", ov.get("search_volume_360d")),
        ("Рост объёма поиска (180д)", _pct(ov.get("search_volume_growth_180d"))),
        ("Кол-во топ товаров", ov.get("num_top_clicked_products")),
        ("Средняя цена (360д)", ov.get("avg_price_360d")),
        ("Диапазон продаж (360д)", ov.get("units_sold_range")),
        ("Уровень возвратов (360д)", _pct(ov.get("return_rate_360d"))),
    ]

    _h(ws, 1, 1, "Метрика")
    _h(ws, 1, 2, "Значение")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25

    for i, (label, value) in enumerate(rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        _c(ws, i, 1, label, fill=fill, bold=True)
        _c(ws, i, 2, value, fill=fill, align="right")


# ── Sheet 2: Competition ───────────────────────────────────────────────────────

def _sheet_competition(wb, schema):
    ws = wb.create_sheet("Competition")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    comp = {}
    ct = schema.get("competition_table")
    if isinstance(ct, dict):
        comp = ct

    METRICS = [
        ("Товары и поиск", "product_count", "Кол-во товаров", "product_and_search"),
        ("Товары и поиск", "sponsored_product_count", "Спонсируемые товары", "product_and_search"),
        ("Товары и поиск", "prime_product_count", "Prime товары", "product_and_search"),
        ("Товары и поиск", "top5_products_click_share", "Доля кликов топ-5 (%)", "product_and_search"),
        ("Товары и поиск", "top20_products_click_share", "Доля кликов топ-20 (%)", "product_and_search"),
        ("Товары и поиск", "avg_selling_price", "Средняя цена продажи", "product_and_search"),
        ("Товары и поиск", "search_volume", "Объём поиска", "product_and_search"),
        ("Товары и поиск", "search_conversion_rate", "Конверсия поиска (%)", "product_and_search"),
        ("Товары и поиск", "new_product_count", "Новые товары", "product_and_search"),
        ("Товары и поиск", "success_launch_product_count", "Успешные запуски", "product_and_search"),
        ("Бренды", "brand_count", "Кол-во брендов", "brands_and_selling_partners"),
        ("Бренды", "top5_brands_click_share", "Доля кликов топ-5 брендов (%)", "brands_and_selling_partners"),
        ("Бренды", "top20_brands_click_share", "Доля кликов топ-20 брендов (%)", "brands_and_selling_partners"),
        ("Бренды", "avg_age_brands_days", "Средний возраст брендов (дней)", "brands_and_selling_partners"),
        ("Бренды", "selling_partner_count", "Кол-во продавцов", "brands_and_selling_partners"),
        ("Покупатели", "avg_rating", "Средний рейтинг", "customer_experience"),
        ("Покупатели", "avg_out_of_stock_rate", "Среднее отсутствие на складе (%)", "customer_experience"),
        ("Покупатели", "avg_bsr", "Средний BSR", "customer_experience"),
        ("Покупатели", "avg_review_count", "Среднее кол-во отзывов", "customer_experience"),
    ]

    _h(ws, 1, 1, "Категория")
    _h(ws, 1, 2, "Метрика")
    _h(ws, 1, 3, "Сегодня")
    _h(ws, 1, 4, "90 дней назад")
    _h(ws, 1, 5, "360 дней назад")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    for i, (category, key, label, section) in enumerate(METRICS, start=2):
        section_data = comp.get(section, {})
        vals = section_data.get(key, {}) if isinstance(section_data, dict) else {}
        if not isinstance(vals, dict):
            vals = {}
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        _c(ws, i, 1, category, fill=fill)
        _c(ws, i, 2, label, fill=fill, bold=True)
        _c(ws, i, 3, vals.get("today"), fill=fill, align="right")
        _c(ws, i, 4, vals.get("90d"), fill=fill, align="right")
        _c(ws, i, 5, vals.get("360d"), fill=fill, align="right")

    ws.auto_filter.ref = f"A1:E{len(METRICS) + 1}"


# ── Sheet 3: Search Terms ──────────────────────────────────────────────────────

def _sheet_search_terms(wb, schema):
    ws = wb.create_sheet("Search Terms")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    terms = []
    st = schema.get("search_terms_csv")
    if isinstance(st, dict):
        terms = st.get("terms") or []
    elif schema.get("chrome_ext") and isinstance(schema["chrome_ext"], dict):
        terms = schema["chrome_ext"].get("search_terms") or []

    headers = ["Запрос", "Объём (360д)", "Рост (90д)", "Рост (180д)",
               "Доля кликов", "Конверсия", "Momentum", "Кластер",
               "#1 Продукт", "#1 ASIN", "#2 Продукт", "#2 ASIN", "#3 Продукт", "#3 ASIN"]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    widths = [35, 14, 12, 12, 14, 12, 12, 20, 40, 14, 40, 14, 40, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_i, term in enumerate(terms, start=2):
        if not isinstance(term, dict):
            continue
        fill = ALT_FILL if row_i % 2 == 0 else WHITE_FILL
        top = term.get("top_products") or []

        def tp(n, field):
            if len(top) > n and isinstance(top[n], dict):
                return top[n].get(field, "")
            return ""

        _c(ws, row_i, 1, term.get("term"), fill=fill, bold=True)
        _c(ws, row_i, 2, term.get("volume_360d"), fill=fill, align="right")
        _c(ws, row_i, 3, _pct(term.get("growth_90d")), fill=fill, align="right")
        _c(ws, row_i, 4, _pct(term.get("growth_180d")), fill=fill, align="right")
        _c(ws, row_i, 5, _pct(term.get("click_share_360d")), fill=fill, align="right")
        _c(ws, row_i, 6, _pct(term.get("conversion_rate_360d")), fill=fill, align="right")
        _c(ws, row_i, 7, term.get("momentum"), fill=fill, align="center")
        _c(ws, row_i, 8, term.get("intent_cluster"), fill=fill)
        _c(ws, row_i, 9, tp(0, "title"), fill=fill)
        _c(ws, row_i, 10, tp(0, "asin"), fill=fill)
        _c(ws, row_i, 11, tp(1, "title"), fill=fill)
        _c(ws, row_i, 12, tp(1, "asin"), fill=fill)
        _c(ws, row_i, 13, tp(2, "title"), fill=fill)
        _c(ws, row_i, 14, tp(2, "asin"), fill=fill)

    if terms:
        ws.auto_filter.ref = f"A1:N{len(terms) + 1}"


# ── Sheet 4: Products ──────────────────────────────────────────────────────────

def _sheet_products(wb, schema):
    ws = wb.create_sheet("Products")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    products = []
    pc = schema.get("products_csv")
    if isinstance(pc, dict):
        products = pc.get("products") or []
    elif schema.get("chrome_ext") and isinstance(schema["chrome_ext"], dict):
        products = schema["chrome_ext"].get("products") or []

    headers = ["#", "ASIN", "Бренд", "Тип", "Доля кликов", "Цена",
               "Отзывы", "Рейтинг", "BSR", "Продавцы", "Отзывов/мес", "Флаг", "Название"]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)

    widths = [5, 14, 18, 10, 13, 9, 10, 9, 10, 10, 13, 35, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_i, p in enumerate(products, start=2):
        if not isinstance(p, dict):
            continue
        ptype = p.get("product_type", "direct")
        fill = WARN_FILL if ptype in ("adjacent", "unrelated") else (ALT_FILL if row_i % 2 == 0 else WHITE_FILL)
        _c(ws, row_i, 1, p.get("rank_by_clicks"), fill=fill, align="center")
        _c(ws, row_i, 2, p.get("asin"), fill=fill)
        _c(ws, row_i, 3, p.get("brand"), fill=fill)
        _c(ws, row_i, 4, ptype, fill=fill, align="center")
        _c(ws, row_i, 5, _pct(p.get("click_share_360d")), fill=fill, align="right")
        _c(ws, row_i, 6, p.get("avg_price_360d"), fill=fill, align="right")
        _c(ws, row_i, 7, p.get("total_ratings"), fill=fill, align="right")
        _c(ws, row_i, 8, p.get("avg_rating"), fill=fill, align="right")
        _c(ws, row_i, 9, p.get("avg_bsr"), fill=fill, align="right")
        _c(ws, row_i, 10, p.get("avg_sellers_count"), fill=fill, align="right")
        _c(ws, row_i, 11, p.get("review_velocity_per_month"), fill=fill, align="right")
        _c(ws, row_i, 12, p.get("velocity_flag") or p.get("product_type_note") or "", fill=fill)
        _c(ws, row_i, 13, p.get("product_name"), fill=fill)

    if products:
        ws.auto_filter.ref = f"A1:M{len(products) + 1}"


# ── Sheet 5: Topic Impact ──────────────────────────────────────────────────────

def _sheet_topic_impact(wb, schema):
    ws = wb.create_sheet("Topic Impact")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    headers = ["Тип", "Топик", "Влияние топ-25%", "Влияние все", "Разрыв", "Тренд топ-25%", "Тренд все"]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)
    widths = [10, 35, 16, 16, 12, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_i = 2
    for tab_key, tab_label, fill in [
        ("topic_impact_positive", "Позитив", POSITIVE_FILL),
        ("topic_impact_negative", "Негатив", NEGATIVE_FILL),
    ]:
        data = schema.get(tab_key)
        if not isinstance(data, dict):
            continue
        impact = data.get("impact_chart", {})
        trend = data.get("trend_chart", {})
        topics = impact.get("topics") or []
        for t in topics:
            if not isinstance(t, dict):
                continue
            _c(ws, row_i, 1, tab_label, fill=fill, bold=True, align="center")
            _c(ws, row_i, 2, t.get("topic"), fill=fill)
            _c(ws, row_i, 3, t.get("top25_impact"), fill=fill, align="right")
            _c(ws, row_i, 4, t.get("all_products_impact"), fill=fill, align="right")
            _c(ws, row_i, 5, t.get("gap"), fill=fill, align="right")
            _c(ws, row_i, 6, trend.get("top25_trend") if t.get("topic") == trend.get("topic_shown") else "", fill=fill, align="center")
            _c(ws, row_i, 7, trend.get("all_products_trend") if t.get("topic") == trend.get("topic_shown") else "", fill=fill, align="center")
            row_i += 1

    if row_i > 2:
        ws.auto_filter.ref = f"A1:G{row_i - 1}"


# ── Sheet 6: Returns ───────────────────────────────────────────────────────────

def _sheet_returns(wb, schema):
    ws = wb.create_sheet("Returns")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    topics = []
    ri = schema.get("returns_insights")
    if isinstance(ri, dict):
        topics = ri.get("return_topics") or []
    elif schema.get("chrome_ext") and isinstance(schema["chrome_ext"], dict):
        topics = schema["chrome_ext"].get("return_topics") or []

    _h(ws, 1, 1, "Топик возврата")
    _h(ws, 1, 2, "% Возвратов")
    _h(ws, 1, 3, "Серьёзность")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    for row_i, t in enumerate(topics, start=2):
        if not isinstance(t, dict):
            continue
        sev = t.get("severity", "")
        fill = NEGATIVE_FILL if sev == "critical" else WARN_FILL if sev == "major" else (ALT_FILL if row_i % 2 == 0 else WHITE_FILL)
        _c(ws, row_i, 1, t.get("topic"), fill=fill, bold=(sev == "critical"))
        _c(ws, row_i, 2, _pct(t.get("return_mention_rate")), fill=fill, align="right")
        _c(ws, row_i, 3, sev, fill=fill, align="center")

    if topics:
        ws.auto_filter.ref = f"A1:C{len(topics) + 1}"


# ── Sheet 7: Analysis Report ───────────────────────────────────────────────────

def _sheet_analysis(wb, markdown_report: str):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120

    _h(ws, 1, 1, "Аналитический отчёт (Markdown)")
    ws.row_dimensions[1].height = 24

    for row_i, line in enumerate(markdown_report.splitlines(), start=2):
        cell = ws.cell(row=row_i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=False)
        if line.startswith("## "):
            cell.font = Font(bold=True, size=13)
        elif line.startswith("### "):
            cell.font = Font(bold=True, size=11)
